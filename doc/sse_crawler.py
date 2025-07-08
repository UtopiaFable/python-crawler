import json
import os
import time
import random
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from utils import safe_request, process_feedback, feedback_to_excel, sort_excel

# test
def merge_cell():
    sort_excel('sse')
    wb = load_workbook(f'./result_sse.xlsx')
    ws = wb.active
    unique = {}
    ws.delete_cols(8)
    for i, row in enumerate(ws.iter_rows()):
        if unique.get(row[0].value):
            unique[row[0].value][1] = i + 1
        else:
            unique[row[0].value] = [i + 1, i + 1]
    for k, (r1, r2) in unique.items():
        for col in 'ABEFGHIJK':
            ws.merge_cells(f'{col}{r1}:{col}{r2}')
    wb.save(f'./上交所反馈意见.xlsx')
    wb.close()

def crawl_sse(begin_page=1, end_page=5, status='2'):
    if os.path.exists('./result_sse.xlsx'):
        current_result = pd.read_excel('./result_sse.xlsx', header=0, index_col=0)
    else:
        current_result = None
    bond_type_map = {
        '0': '小公募',
        '1': '私募',
        '2': 'ABS',
        '3': '大公募',
        '4': '基础设施公募REITs',
        '5': '小公募',
        '6': '私募',
        '7': '大公募',
    }
    audit_status_map = {
        '0': '已申报',
        '1': '已受理',
        '2': '已反馈',
        '3': '通过',
        '4': '未通过',
        '11': '提交注册',
        '12': '注册生效',
        '8': '终止',
    }
    index = "https://query.sse.com.cn/commonSoaQuery.do"
    doc_download='https://static.sse.com.cn/bond/'
    data = {
        'jsonCallBack': f'jsonpCallback{random.randint(10000, 99999)}',
        'isPagination': 'true',
        'sqlId': 'ZQ_XMLB',
        'pageHelp.pageSize': '20',
        'status': status,
        'bond_type': '0,5', # '0,5': 小公募 '1,6': 私募 '2': ABS '3,7': 大公募 '4': 基础设施公募
    }
    page_data = {
        'jsonCallBack': f'jsonpCallback{random.randint(10000000, 99999999)}',
        'isPagination': 'false',
    }
    for i in range(begin_page, end_page + 1):
        print(f"***************** page {i} *****************")
        data['pageHelp.pageNo'] = str(i)
        reports = safe_request(index, params=data, headers={'Referer': 'https://www.sse.com.cn/'})
        page = json.loads(reports.content.decode().split('(', maxsplit=1)[-1][:-1])
        for bond in page['pageHelp']['data']:
            bond_num = bond['BOND_NUM']
            bond_name = bond['AUDIT_NAME']
            update_time = bond['PUBLISH_DATE']
            state = audit_status_map[bond['AUDIT_STATUS']]
            print(f'processing {bond_name}, state: {state}, update_time: {update_time}')
            if current_result is not None and bond_name in current_result.index:
                print('\tThis bond has been processed. Skip.\n')
                continue
            year = update_time.split('-')[0]
            page_data['audit_id'] = bond_num
            page_data['sqlId'] = 'ZQ_GGJG'
            reports = safe_request(index, params=page_data, headers={'Referer': 'https://www.sse.com.cn/'})
            bond_file = json.loads(reports.content.decode().split('(', maxsplit=1)[1][:-1])
            doc_feedback = {}
            for feedback in bond_file['result']:
                if feedback['MAIN_TYPE'] != '':
                    continue
                doc_name = feedback['FILE_TITLE']
                doc_url = doc_download + feedback['FILE_PATH']
                print(f'\tprocessing {doc_name}', end='......')
                file_name, suffix = os.path.splitext(doc_name)
                if '反馈意见' not in file_name or '回复' in file_name:
                    print('skip')
                    continue
                max_retry = 5
                while max_retry:
                    try:
                        doc_string = safe_request(doc_url)
                        doc_io = BytesIO(doc_string.content)
                        doc_feedback[feedback['FILE_TIME']] = process_feedback(doc_io, suffix)
                        doc_io.close()
                        print('done')
                        break
                    except:
                        max_retry -= 1
                        if max_retry == 0:
                            doc_feedback[feedback['FILE_TIME']] = '文件处理失败'
                            print('failed')
                        else:
                            time.sleep(0.5)
            if not doc_feedback:
                print('\tNo feedback document found.')
            bond_info = {
                'name': bond_name,
                'state': state,
                'update_date': update_time,
                'feedback': doc_feedback,
                'bond_type': bond_type_map[bond['BOND_TYPE']],
                'scale': bond['PLAN_ISSUE_AMOUNT'],
                'issuer': bond['FULL_NAME'],
                'area': '',
                'underwriter': bond['WRITER_NAME'],
                'file_id': bond['WEN_HAO'],
                'accept_date': bond['ACCEPT_DATE'],
            }
            feedback_to_excel(bond_info, 'sse')
            print()
            time.sleep(1)


if os.path.exists('./result_sse.xlsx'):
    os.remove('./result_sse.xlsx')
for status in ['2', '3', '11', '12', '8']:
    crawl_sse(end_page=5, status=status)
    break
merge_cell()