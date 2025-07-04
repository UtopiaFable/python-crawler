import os
import docx
import time
import pandas as pd
import requests as rq
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from win32com import client

def process_doc(doc_name):
    assert doc_name.endswith(('.doc', '.docx')), "反馈意见文件格式非doc或docx！"
    if doc_name.endswith('doc'):
        word = client.Dispatch("Word.Application")
        abs_doc_path = os.path.abspath(f'../files/{doc_name}')
        doc = word.Documents.Open(abs_doc_path)
        doc.SaveAs(f"{abs_doc_path}x", 12)  # 12代表docx格式
        doc.Close()
        os.remove(abs_doc_path)
        doc_name = doc_name + 'x'
    with open(f'../files/{doc_name}', 'rb') as f:
        doc = docx.Document(f)
        texts = []
        for paragraph in doc.paragraphs:
            texts.append(paragraph.text)
    start, end = 0, -1
    for i in range(len(texts)):
        if texts[i].endswith('反馈意见：'):
            start = i
        elif texts[i].find('请你公司对上述问题逐项落实') != -1:
            end = i
    feedback = '\n'.join(texts[start+1:end]).strip()
    return feedback

def feedback_to_excel(bond_info):
    if os.path.exists('./result.xlsx'):
        df = pd.read_excel('./result.xlsx')
    else:
        df = pd.DataFrame({
            '债券名': [],
            '状态': [],
            '反馈意见': [],
            '反馈时间': [],
            '债券类别': [],
            '申报规模（亿元）': [],
            '发行人': [],
            '地区': [],
            '承销商/管理人': [],
            '文件文号': [],
            '受理日期': [],
            '更新日期': [],
        })
    feedback = bond_info['feedback'] or {'N/A': 'N/A'}
    for date, text in feedback.items():
        new_row = pd.DataFrame({
            '债券名': [bond_info['name']],
            '状态': [bond_info['state']],
            '反馈意见': [text],
            '反馈时间': [date],
            '债券类别': [bond_info['bond_type']],
            '申报规模（亿元）': [bond_info['scale']],
            '发行人': [bond_info['issuer']],
            '地区': [bond_info['area']],
            '承销商/管理人': [bond_info['underwriter']],
            '文件文号': [bond_info['file_id']],
            '受理日期': [bond_info['accept_date']],
            '更新日期': [bond_info['update_date']],
        })
        df = pd.concat([df, new_row])
    df.to_excel('./result.xlsx', index=False)

def merge_cell():
    wb = load_workbook('./doc/result.xlsx')
    ws = wb.active
    # align = Alignment(horizontal='general', vertical='center', wrap_text=True)
    unique = {}
    for i, row in enumerate(ws.iter_rows()):
        if unique.get(row[0].value):
            unique[row[0].value][1] = i + 1
        else:
            unique[row[0].value] = (i + 1, i + 1)
    for k, (r1, r2) in unique.items():
        for col in 'ABEFGHIJKL':
            ws.merge_cells(f'{col}{r1}:{col}{r2}')
    wb.save('./result.xlsx')
    wb.close()


def crawl_szse():
    index = "https://bond.szse.cn/api/report/ShowReport/data"
    doc_download='https://reportdocs.static.szse.cn'
    data = {
        'SHOWTYPE': 'JSON',
        'CATALOGID': 'xmjdxx',
        'TABKEY': 'tab1',
        'zqlb': '0', # 0: 小公募 1: 私募 2: ABS 3: 大公募
        'selectXmztt': '已反馈,通过,提交注册,注册生效,终止'
    }
    i = 0
    year = 2025
    if os.path.exists('./result.xlsx'):
        os.remove('./result.xlsx')
    while i < 5:
        i += 1
        print(f"***************** page {i} *****************")
        data['PAGENO'] = str(i)
        while True:
            try:
                reports = rq.get(index,params=data)
            except:
                time.sleep(2)
                continue
            time.sleep(0.5)
            break
        for bond in reports.json()[0]['data']:
            soup = BeautifulSoup(bond['zqmc'], 'html.parser').find('a')
            bond_url, bond_name = soup.get('a-param'), soup.string
            update_time = bond['xmztgxrq']
            state = bond['xmzt']
            print(f'processing {bond_name}, state: {state}, update_time: {update_time}')
            year = update_time
            while True:
                try:
                    bond_page = rq.get(index + '?' + bond_url.split('?')[1])
                except:
                    time.sleep(2)
                    continue
                time.sleep(0.5)
                break
            doc_feedback = {}
            for feedback in bond_page.json()[2]['data']:
                soup = BeautifulSoup(feedback['fkyjh'], 'html.parser').find('a')
                doc_url, doc_name = soup.get('encode-open'), soup.string
                print(f'\tprocessing {doc_name}', end='......')
                if not os.path.splitext(doc_name)[0].endswith('反馈意见'):
                    print('skip')
                    continue
                while True:
                    try:
                        doc_string = rq.get(doc_download + doc_url)
                    except:
                        time.sleep(2)
                        continue
                    time.sleep(0.5)
                    break
                with open(f'../files/{doc_name}', 'wb') as f:
                    f.write(doc_string.content)
                while not os.path.exists(f'../files/{doc_name}'):
                    time.sleep(1)
                doc_feedback[feedback['fkyjhgxrq']] = process_doc(doc_name)
                if doc_name.endswith('.doc'):
                    doc_name = doc_name + 'x'
                os.remove(f'../files/{doc_name}')
                print('done')
            if not doc_feedback:
                print('\tNo feedback document found.')
            assert len(bond_page.json()[0]['data']) == 1, '项目基本信息不唯一！'
            basic_info = bond_page.json()[0]['data'][0]
            bond_info = {
                'name': bond_name,
                'state': state,
                'update_date': update_time,
                'feedback': doc_feedback,
                'bond_type': basic_info['zqlb'],
                'scale': basic_info['nfxje'],
                'issuer': basic_info['fxr'],
                'area': basic_info['dq'],
                'underwriter': basic_info['cxsqc'],
                'file_id': basic_info['jysqrwjwh'],
                'accept_date': bond['xmslrq'],
            }
            feedback_to_excel(bond_info)
            print()
            time.sleep(1)

crawl_szse()
merge_cell()