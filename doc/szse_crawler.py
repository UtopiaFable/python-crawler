import os
import time
from io import BytesIO

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from utils import safe_request, process_feedback, feedback_to_excel, sort_excel

def merge_cell(suffix='szse'):
    sort_excel('szse')
    wb = load_workbook(f'./result_{suffix}.xlsx')
    ws = wb.active
    # align = Alignment(horizontal='general', vertical='center', wrap_text=True)
    unique = {}
    for i, row in enumerate(ws.iter_rows()):
        if unique.get(row[0].value):
            unique[row[0].value][1] = i + 1
        else:
            unique[row[0].value] = [i + 1, i + 1]
    for k, (r1, r2) in unique.items():
        for col in 'ABEFGHIJKL':
            ws.merge_cells(f'{col}{r1}:{col}{r2}')
    wb.save(f'./深交所反馈意见.xlsx')
    wb.close()


def crawl_szse(begin_page=1, end_page=5):
    if os.path.exists('./result_sse.xlsx'):
        current_result = pd.read_excel('./result_szse.xlsx', header=0, index_col=0)
    else:
        current_result = None
    index = "https://bond.szse.cn/api/report/ShowReport/data"
    doc_download='https://reportdocs.static.szse.cn'
    data = {
        'SHOWTYPE': 'JSON',
        'CATALOGID': 'xmjdxx',
        'TABKEY': 'tab1',
        'zqlb': '0', # 0: 小公募 1: 私募 2: ABS 3: 大公募
        'selectXmztt': '已反馈,通过,提交注册,注册生效,终止'
    }
    for i in range(begin_page, end_page + 1):
        print(f"***************** page {i} *****************")
        data['PAGENO'] = str(i)
        reports = safe_request(index,params=data)
        for bond in reports.json()[0]['data']:
            soup = BeautifulSoup(bond['zqmc'], 'html.parser').find('a')
            bond_url, bond_name = soup.get('a-param'), soup.string
            update_time = bond['xmztgxrq']
            state = bond['xmzt']
            print(f'processing {bond_name}, state: {state}, update_time: {update_time}')
            if current_result is not None and bond_name in current_result.index:
                print('\tThis bond has been processed. Skip.\n')
                continue
            year = update_time.split('-')[0]
            bond_page = safe_request(index + '?' + bond_url.split('?')[1])
            doc_feedback = {}
            for feedback in bond_page.json()[2]['data']:
                soup = BeautifulSoup(feedback['fkyjh'], 'html.parser').find('a')
                doc_url, doc_name = soup.get('encode-open'), soup.string
                print(f'\tprocessing {doc_name}', end='......')
                file_name, suffix = os.path.splitext(doc_name)
                if '反馈意见' not in file_name or '回复' in file_name:
                    print('skip')
                    continue
                max_retry = 5
                while max_retry:
                    try:
                        doc_string = safe_request(doc_download + doc_url)
                        doc_io = BytesIO(doc_string.content)
                        doc_feedback[feedback['fkyjhgxrq']] = process_feedback(doc_io, suffix)
                        doc_io.close()
                        break
                    except:
                        max_retry -= 1
                        time.sleep(0.5)
                print('done')
            if not doc_feedback:
                print('\tNo feedback document found.')
            assert len(bond_page.json()[0]['data']) == 1, '项目基本信息不唯一！'
            basic_info = bond_page.json()[0]['data'][0]
            bond_info = {
                'name': bond_name,
                'state': state,
                'update_date': update_time,
                'feedback': doc_feedback,
                'bond_type': basic_info['zqlb'],
                'scale': basic_info['nfxje'],
                'issuer': basic_info['fxr'],
                'area': basic_info['dq'],
                'underwriter': basic_info['cxsqc'],
                'file_id': basic_info['jysqrwjwh'],
                'accept_date': bond['xmslrq'],
            }
            feedback_to_excel(bond_info, 'szse')
            print()
            time.sleep(1)

if os.path.exists('./result_szse.xlsx'):
    os.remove('./result_szse.xlsx')
crawl_szse()
merge_cell()